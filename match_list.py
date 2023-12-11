import pandas as pd
from openpyxl import load_workbook, Workbook
import os

class MatchList:
    
    def __init__(self, path = 'match_list.xlsx') -> None:
        self.PATH = path
        self.BOOK : Workbook
        self.load_file()
    
    def add_match(self, date : str,
                  green_atk : str, green_def : str,
                  yellow_atk : str, yellow_def : str,
                  green_score : int, yellow_score : int,
                  added_by : str):
        
        ws = self.load_file()
        
        data = [date, green_atk, green_def,
                yellow_atk, yellow_def,
                green_score, yellow_score, added_by]
        
        ws.append(data)
        self.save_file()

    def last_match(self):
        self.load_file()
        match_list = list(self.BOOK.active.iter_rows(values_only=True))
        return match_list[-1]

    def remove_last_match(self):
        self.load_file()
        to_return = self.last_match()
        
        self.BOOK.active.delete_rows(self.BOOK.active.max_row)
        self.save_file()
            
        return to_return
    
    def save_file(self):
        self.BOOK.save(self.PATH)
    
    def load_file(self):
        if os.path.exists(self.PATH):
            self.BOOK = load_workbook(filename = self.PATH)
            self.BOOK.active
            return self.BOOK.active
        else:
            self.BOOK = Workbook()
            header = ["Date", "Green ATK", "Green DEF",
            "Yellow ATK", "Yellow DEF",
            "Green Score", "Yellow Score", "Added By"]
            self.BOOK.active.append(header)
            self.save_file()

if __name__ == '__main__':
    from icecream import ic
    from datetime import datetime
    
    current_time = datetime.now()
    formatted_time = current_time.strftime('%Y-%m-%d %H:%M')
    print(formatted_time)
    
    match_list = MatchList()
    match_list.add_match(formatted_time,
                         "Niki Di Giano",
                         "Pasquale Barbato",
                         "Giulio Gualandi",
                         "Ciro Pentangelo",
                         7, 8, "Niki")
    
    match_list.add_match(formatted_time,
                         "Niki Di Giano",
                         "Pasquale Barbato",
                         "Vittorio Grimaldi",
                         "Ciro Pentangelo",
                         7, 8, "Niki")
    
    data = match_list.load_file()
    ic(data)
    latest_match = match_list.remove_last_match()